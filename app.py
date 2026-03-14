import streamlit as st
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
import io

st.title("Attendance Dashboard")

uploaded_file = st.file_uploader("Upload Attendance PDF", type="pdf")

if uploaded_file:

    rows = []

    with pdfplumber.open(uploaded_file) as pdf:
        for page in pdf.pages:
            table = page.extract_table()

            if table:
                for row in table[1:]:
                    rows.append(row)

    df = pd.DataFrame(rows)

    df = df[[1,5]]
    df.columns = ["Course Name","Attendance"]
    df = df.dropna()

    total = df.groupby("Course Name").size()
    present = df[df["Attendance"]=="P"].groupby("Course Name").size()

    summary = pd.DataFrame({
        "Total": total,
        "Present": present
    }).fillna(0)

    wb = load_workbook("for gpt.xlsx")
    ws = wb.active

    for row in range(2,30):

        subject = ws.cell(row=row,column=2).value

        if subject in summary.index:
            ws.cell(row=row,column=3).value = int(summary.loc[subject,"Total"])
            ws.cell(row=row,column=4).value = int(summary.loc[subject,"Present"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    result_df = pd.read_excel(output)

    st.subheader("Attendance Table")

    st.dataframe(result_df)
